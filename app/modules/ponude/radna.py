"""Uvoz "radne tablice" (format kupca) u upit + logistika/DAP izračun.

Radna tablica po stavci nosi: broj ERP kalkulacije (veza na našu arhivu!),
kupčev CODE, tehničke podatke za excel ponudu, order quantity (KOMADI) i
pakiranje (kom/kutiji, kutija/paleti) za izračun broja paleta.

DAP model (potvrdio korisnik):
  cijena po paleti  = ukupni prijevoz / ukupan broj paleta u upitu
  prijevoz stavke   = palete stavke × cijena po paleti
  prijevoz €/1000   = prijevoz stavke / (količina kom / 1000)
  DAP €/1000        = EXW €/1000 (iz kalkulacije + marža) + prijevoz €/1000
"""
import json

import openpyxl
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.modules.arhiva import models as am
from app.modules.ponude import models as pm

# header (lowercase, startswith) -> interni ključ
_MAPA = [
    ("kalkulacija", "kalk_broj"),
    ("novi alat", "novi_alat"),
    ("code", "code"),
    ("description", "description"),
    ("size", "size_cl"),
    ("label location", "location"),
    ("width", "width"),
    ("height", "height"),
    ("material", "material"),
    ("embossing", "embossing"),
    ("gsm", "gsm"),
    ("print", "print_type"),
    ("# colours", "colours"),
    ("varnish", "varnish"),
    ("cutting", "cutting"),
    ("order quantity", "qty"),
    ("kom/kutiji", "kom_kutija"),
    ("kutija na paleti", "kutija_paleta"),
    ("cijena alata", "alat"),
]

_ERR = {"#VALUE!", "#REF!", "#N/A", "#DIV/0!", "#NAME?"}


def _cist(v):
    if v is None:
        return None
    if isinstance(v, float) and v.is_integer():
        return int(v)
    s = str(v).strip()
    return None if (not s or s in _ERR) else v if not isinstance(v, str) else s


def _broj(v):
    v = _cist(v)
    if v is None or isinstance(v, (int, float)):
        return v
    try:
        return float(str(v).replace(",", "."))
    except ValueError:
        return None


def parsiraj_radnu(putanja: str) -> list[dict]:
    """Vrati listu dict-ova stavki iz radne tablice (prvi list s 'CODE' headerom)."""
    wb = openpyxl.load_workbook(putanja, data_only=True)
    for ws in wb.worksheets:
        # nađi header red: sadrži ćeliju koja počinje s "CODE"
        for r in range(1, min(ws.max_row, 60) + 1):
            stupci = {}
            for c in range(1, ws.max_column + 1):
                v = ws.cell(row=r, column=c).value
                if isinstance(v, str):
                    h = v.strip().lower()
                    for prefiks, kljuc in _MAPA:
                        if h.startswith(prefiks) and kljuc not in stupci:
                            stupci[kljuc] = c
            if "code" in stupci and "kalk_broj" in stupci:
                return _procitaj_redove(ws, r, stupci)
    raise ValueError("Nije pronađen header radne tablice (kolone CODE + KALKULACIJA).")


def _procitaj_redove(ws, r_hdr: int, stupci: dict) -> list[dict]:
    redovi, praznih = [], 0
    r = r_hdr + 1
    while r <= ws.max_row and praznih < 5:
        red = {k: _cist(ws.cell(row=r, column=c).value) for k, c in stupci.items()}
        if red.get("code") or red.get("kalk_broj"):
            praznih = 0
            for k in ("qty", "kom_kutija", "kutija_paleta", "alat",
                      "width", "height", "size_cl", "colours", "gsm"):
                red[k] = _broj(red.get(k))
            redovi.append(red)
        else:
            praznih += 1
        r += 1
    return redovi


def uvezi_u_upit(db: Session, upit: pm.Upit, redovi: list[dict]) -> list[dict]:
    """Kreiraj stavke upita iz redova radne tablice. Vraća izvještaj po redu."""
    izvjestaj = []
    postojece = {s.kalkulacija_id for s in upit.stavke}
    for red in redovi:
        oznaka = str(red.get("code") or red.get("kalk_broj") or "?")
        broj = str(red.get("kalk_broj") or "").strip()
        k = db.scalar(select(am.Kalkulacija).where(am.Kalkulacija.broj == broj)) if broj else None
        if k is None:
            izvjestaj.append({"oznaka": oznaka, "status": "greska",
                              "poruka": f"Kalkulacija {broj or '—'} nije u arhivi — "
                                        f"prvo uvezi Excel kalkulaciju."})
            continue
        if k.id in postojece:
            izvjestaj.append({"oznaka": oznaka, "status": "preskoceno",
                              "poruka": f"Kalkulacija {broj} je već u upitu."})
            continue
        tehnika = {kk: red.get(kk) for kk in
                   ("code", "description", "size_cl", "location", "width", "height",
                    "material", "embossing", "gsm", "print_type", "colours",
                    "varnish", "cutting")}
        qty = red.get("qty")
        db.add(pm.UpitStavka(
            upit_id=upit.id, proizvod_id=k.proizvod_id, kalkulacija_id=k.id,
            kolicina=(qty / 1000.0) if qty else k.serija, marza_pct=30.0,
            kom_kutija=red.get("kom_kutija"), kutija_paleta=red.get("kutija_paleta"),
            alat_cijena=red.get("alat"),
            tehnika=json.dumps(tehnika, ensure_ascii=False),
        ))
        postojece.add(k.id)
        izvjestaj.append({"oznaka": oznaka, "status": "ok",
                          "poruka": f"Dodano — kalkulacija {broj}, "
                                    f"{qty:,.0f} kom".replace(",", ".") if qty else
                                    f"Dodano — kalkulacija {broj}"})
    db.commit()
    return izvjestaj


def logistika(upit: pm.Upit) -> dict:
    """Izračun paleta, prijevoza i DAP-a po stavci (živo, iz trenutnih podataka)."""
    po_stavci: dict[int, dict] = {}
    ukupno_paleta = 0
    for s in upit.stavke:
        p = s.palete
        if p:
            ukupno_paleta += p
    po_paleti = (upit.prijevoz_ukupno / ukupno_paleta) \
        if (upit.prijevoz_ukupno and ukupno_paleta) else None
    for s in upit.stavke:
        d = {"palete": s.palete, "exw_1000": s.exw_1000,
             "prijevoz_stavka": None, "prijevoz_1000": None, "dap_1000": None}
        if s.palete and po_paleti is not None and s.kolicina_kom:
            d["prijevoz_stavka"] = round(s.palete * po_paleti, 2)
            d["prijevoz_1000"] = round(d["prijevoz_stavka"] / (s.kolicina_kom / 1000.0), 4)
            if s.exw_1000 is not None:
                d["dap_1000"] = round(s.exw_1000 + d["prijevoz_1000"], 4)
        elif s.exw_1000 is not None:
            d["dap_1000"] = None  # nema pakiranja/prijevoza -> DAP se ne može izračunati
        po_stavci[s.id] = d
    return {"po_stavci": po_stavci, "ukupno_paleta": ukupno_paleta,
            "po_paleti": round(po_paleti, 2) if po_paleti else None}
