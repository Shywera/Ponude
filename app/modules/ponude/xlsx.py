"""XLSX export ponude — "excel" stil za velike kupce, po uzoru na
Resources/Offer 1_2 Brasserie Champigneulles: logo, OFFER br., tablica
stavki s tehničkim kolonama + cijena €/1000 + Total + Tools, suma, Remarks.

Cijena kolona ovisi o paritetu ponude: DAP -> dap_1000, inače EXW -> exw_1000
(jed_cijena/iznos su ionako zamrznuti na odabrani paritet pri generiranju).
"""
import io
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.modules.ponude.models import Ponuda

_LOGO = str(Path(__file__).parent / "logo.png")

_HDR = ["Description", "Size cl", "Label Location",
        "Width (mm) length left to right *read from front",
        "Height (mm) top to bottom *read from front",
        "Material Type", "Embossing", "gsm", "Print Type", "# Colours",
        "Varnish Type(s)", "Cutting Method", "Order quantity",
        None,  # cijena kolona — naslov ovisi o paritetu
        "Total", "Tools (one time price)"]

_TANKA = Side(style="thin", color="9CA3AF")
_RUB = Border(left=_TANKA, right=_TANKA, top=_TANKA, bottom=_TANKA)


def xlsx_ponuda(p: Ponuda) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Offer"

    # logo/zaglavlje (naziv + adresa iz konfiguracije)
    try:
        img = XLImage(_LOGO)
        skala = 120.0 / img.height  # ~120 px visine
        img.height, img.width = int(img.height * skala), int(img.width * skala)
        ws.add_image(img, "A1")
    except Exception:
        from app.core.config import settings as _cfg
        ws["A1"] = f"{_cfg.firma_naziv}, {_cfg.firma_adresa.lstrip('A ').strip()}"

    ws["N2"] = p.kupac_naziv
    ws["N2"].font = Font(bold=True, size=12)
    ws["N3"] = f"Date: {p.datum.strftime('%d.%m.%Y.')}"

    ws["A9"] = f"OFFER {p.broj}"
    ws["A9"].font = Font(bold=True, size=14)

    # naslov cjenovne kolone prema paritetu
    par = (p.paritet or "EXW").strip()
    hdr = list(_HDR)
    hdr[13] = f"{par} (EUR/1000)"

    r_hdr = 11
    fill = PatternFill(start_color="FF1F2937", end_color="FF1F2937", fill_type="solid")
    for c, naslov in enumerate(hdr, start=1):
        cel = ws.cell(row=r_hdr, column=c, value=naslov)
        cel.font = Font(bold=True, color="FFFFFF", size=9)
        cel.fill = fill
        cel.border = _RUB
        cel.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    ws.row_dimensions[r_hdr].height = 42

    r = r_hdr + 1
    for s in p.stavke:
        teh = json.loads(s.tehnika) if s.tehnika else {}
        dap = par.upper().startswith("DAP")
        cijena = (s.dap_1000 if dap else s.exw_1000)
        if cijena is None:
            cijena = s.jed_cijena
        qty_kom = (s.kolicina or 0) * 1000.0
        total = round(qty_kom / 1000.0 * cijena, 2) if cijena is not None else None
        vals = [teh.get("description") or s.naziv, teh.get("size_cl"),
                teh.get("location"), teh.get("width"), teh.get("height"),
                teh.get("material"), teh.get("embossing"), teh.get("gsm"),
                teh.get("print_type"), teh.get("colours"), teh.get("varnish"),
                teh.get("cutting"), qty_kom or None, cijena, total, s.alat_cijena]
        for c, v in enumerate(vals, start=1):
            cel = ws.cell(row=r, column=c, value=v)
            cel.border = _RUB
            cel.font = Font(size=9)
            if c in (13,):
                cel.number_format = "#,##0"
            elif c in (14,):
                cel.number_format = "#,##0.0000"
            elif c in (15, 16):
                cel.number_format = "#,##0.00"
        r += 1

    # suma
    ws.cell(row=r, column=14, value="TOTAL:").font = Font(bold=True, size=10)
    suma = ws.cell(row=r, column=15, value=round(sum(
        (x.value or 0) for x in [ws.cell(row=rr, column=15)
                                 for rr in range(r_hdr + 1, r)]), 2))
    suma.font = Font(bold=True, size=10)
    suma.number_format = "#,##0.00"
    suma_alat = ws.cell(row=r, column=16, value=round(sum(
        (s.alat_cijena or 0) for s in p.stavke), 2) or None)
    if suma_alat.value:
        suma_alat.font = Font(bold=True, size=10)
        suma_alat.number_format = "#,##0.00"

    # napomene
    r += 3
    ws.cell(row=r, column=1, value="Remarks:").font = Font(bold=True)
    if p.napomene:
        for i, linija in enumerate(p.napomene.split("\n")):
            ws.cell(row=r + 1 + i, column=1, value=linija)

    # širine kolona
    sirine = [26, 7, 12, 12, 12, 16, 10, 6, 10, 9, 11, 12, 13, 14, 12, 14]
    for i, w in enumerate(sirine, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = f"A{r_hdr + 1}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
