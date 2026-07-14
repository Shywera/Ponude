"""Parser Excel kalkulacija "Prodajna cijena <broj>.xlsx".

KLJUČNO: raspored NIJE fiksan — broj artikala (1/2/6), materijala i operacija
varira po datoteci pa se svi blokovi pomiču. Parser se zato sidri na tekstualne
oznake ("KALKULACIJA BR", "Proces", "UTROŠAK MATERIJALA", ...), nikad na
apsolutne adrese ćelija.

Troškovi se RAČUNAJU iz sirovih stavki (materijali + operacije rada) — Excelov
keš prodajne cijene se ignorira jer dio izvornih datoteka ima `==` tipfeler u
"Proizvod" bloku koji ruši cijelu kaskadu (223 #VALUE! ćelije po datoteci).
"""
import re
import unicodedata
from dataclasses import dataclass, field
from datetime import date, datetime, time, timedelta

import openpyxl


# ---------------------------------------------------------------- strukture

@dataclass
class ArtiklSpec:
    rb: int | None = None
    naziv: str | None = None
    sirina_mm: float | None = None
    visina_mm: float | None = None
    broj_boja: float | None = None
    broj_prolaza: float | None = None
    lak: str | None = None
    dorada: str | None = None
    tip_papira: str | None = None
    gramatura: str | None = None
    nabavna_cijena: str | None = None
    udio: float | None = None
    kolicina: float | None = None
    ck_total: float | None = None


@dataclass
class Materijal:
    sifra: str | None = None
    naziv: str | None = None
    jedinica: str | None = None
    kolicina: float | None = None
    jed_cijena: float | None = None
    vrijednost: float | None = None


@dataclass
class Operacija:
    stroj: str | None = None
    operacija: str | None = None
    proces: int | None = None
    vrijeme_min: float | None = None
    cijena_h: float | None = None
    iznos: float | None = None


@dataclass
class Kalkulacija:
    broj: str = ""
    normativ_br: str | None = None
    datum: date | None = None
    kupac: str = ""
    naziv_proizvoda: str = ""
    serija: float | None = None
    serija_jedinica: str | None = None
    glavni_stroj: str | None = None

    artikli: list[ArtiklSpec] = field(default_factory=list)
    materijali: list[Materijal] = field(default_factory=list)
    operacije: list[Operacija] = field(default_factory=list)

    # izračunato iz stavki
    trosak_materijal: float = 0.0
    trosak_rad: float = 0.0
    vanjska_usluga: float = 0.0
    ck_total: float = 0.0
    ck_po_jedinici: float | None = None

    upozorenja: list[str] = field(default_factory=list)
    datoteka: str | None = None


# ---------------------------------------------------------------- pomoćne

_ERR_VALUES = {"#VALUE!", "#REF!", "#DIV/0!", "#N/A", "#NAME?", "#NULL!", "#NUM!"}


def _txt(v) -> str | None:
    """Vrijednost ćelije kao očišćeni string (None za prazno/greške)."""
    if v is None:
        return None
    s = str(v).strip()
    if not s or s in _ERR_VALUES:
        return None
    return s


def _num(v) -> float | None:
    """Vrijednost ćelije kao broj (None ako nije broj ili je Excel greška)."""
    if v is None or isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s or s in _ERR_VALUES:
        return None
    try:
        return float(s.replace(",", "."))
    except ValueError:
        return None


def _minute(v) -> float | None:
    """Vrijeme rada u minutama. Podnosi timedelta, time, "HH:MM" (i >24h npr "25:15")."""
    if v is None:
        return None
    if isinstance(v, timedelta):
        return v.total_seconds() / 60.0
    if isinstance(v, time):
        return v.hour * 60.0 + v.minute + v.second / 60.0
    if isinstance(v, datetime):
        return v.hour * 60.0 + v.minute
    if isinstance(v, (int, float)):
        # Excel serijsko vrijeme: 1.0 == 24h
        return float(v) * 24 * 60
    s = str(v).strip()
    m = re.match(r"^(\d+):(\d{1,2})(?::(\d{1,2}))?$", s)
    if m:
        h, mi, se = int(m.group(1)), int(m.group(2)), int(m.group(3) or 0)
        return h * 60.0 + mi + se / 60.0
    return None


def _datum(v) -> date | None:
    """Datum iz ćelije: datetime ili string "dd.mm.yyyy."."""
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = _txt(v)
    if not s:
        return None
    m = re.match(r"^(\d{1,2})\.(\d{1,2})\.(\d{4})\.?$", s)
    if m:
        try:
            return date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
        except ValueError:
            return None
    return None


def normaliziraj_naziv(naziv: str) -> str:
    """Ključ proizvoda iz naziva: mala slova, bez dijakritike, sažeti razmaci.
    Završni brojčani tokeni (količina/serija u nazivu, npr. "UV LAK 6.000" ili
    "... 40000545 850") se skidaju — količinske varijante su ISTI proizvod."""
    s = unicodedata.normalize("NFKD", naziv)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r"[^\w\s]", " ", s.lower())
    tokeni = re.sub(r"\s+", " ", s).strip().split(" ")
    while tokeni and tokeni[-1].isdigit():
        tokeni.pop()
    return " ".join(tokeni)


def izvuci_sifru_artikla(naziv: str) -> str | None:
    """Kupčeva šifra artikla iz naziva proizvoda (zadnji niz od >=6 znamenki)."""
    kandidati = re.findall(r"\d{6,}", naziv)
    return kandidati[-1] if kandidati else None


class _List:
    """Radni omotač oko openpyxl lista: traženje oznaka + čitanje redaka."""

    def __init__(self, ws):
        self.ws = ws
        self.max_row = ws.max_row
        self.max_col = ws.max_column

    def cell(self, r: int, c: int):
        return self.ws.cell(row=r, column=c).value

    def nadji(self, oznaka: str, od_reda: int = 1, stupac: int | None = None) -> tuple[int, int] | None:
        """Prvi redak (od `od_reda`) čija ćelija počinje s `oznaka` (case-insensitive).
        Ako je `stupac` zadan, gleda samo taj stupac."""
        o = oznaka.lower()
        for r in range(od_reda, self.max_row + 1):
            stupci = [stupac] if stupac else range(1, self.max_col + 1)
            for c in stupci:
                v = self.cell(r, c)
                if isinstance(v, str) and v.strip().lower().startswith(o):
                    return (r, c)
        return None

    def desno_od(self, r: int, c: int):
        """Prva neprazna vrijednost desno od (r, c) u istom retku."""
        for cc in range(c + 1, self.max_col + 1):
            v = self.cell(r, cc)
            if v is not None and _txt(v) is not None:
                return v
        return None


# ---------------------------------------------------------------- parser

def parsiraj(putanja: str) -> Kalkulacija:
    """Parsiraj jednu datoteku kalkulacije. Diže ValueError ako format ne odgovara."""
    wb = openpyxl.load_workbook(putanja, data_only=True)
    ws = wb.active
    L = _List(ws)
    k = Kalkulacija(datoteka=str(putanja).replace("\\", "/").split("/")[-1])

    # --- zaglavlje ------------------------------------------------------
    poz = L.nadji("KALKULACIJA BR")
    if poz is None:
        raise ValueError("Nije pronađena oznaka 'KALKULACIJA BR' — nije kalkulacija?")
    r1, c1 = poz
    k.broj = _txt(L.desno_od(r1, c1)) or ""
    if not k.broj:
        raise ValueError("Prazan broj kalkulacije.")

    poz = L.nadji("NORMATIV BR", od_reda=r1)
    if poz:
        k.normativ_br = _txt(L.desno_od(*poz))

    poz = L.nadji("DATUM", od_reda=r1)
    if poz:
        k.datum = _datum(L.desno_od(*poz))

    poz = L.nadji("KUPAC", od_reda=r1)
    if poz:
        k.kupac = _txt(L.desno_od(*poz)) or ""

    poz = L.nadji("NAZIV PROIZVODA", od_reda=r1)
    if poz:
        k.naziv_proizvoda = _txt(L.desno_od(*poz)) or ""

    poz = L.nadji("SERIJA")
    if poz:
        r, c = poz
        k.serija = _num(L.desno_od(r, c))
        # jedinica je u ćeliji nakon količine
        for cc in range(c + 1, L.max_col + 1):
            if _num(L.cell(r, cc)) is not None:
                k.serija_jedinica = _txt(L.cell(r, cc + 1))
                break

    # --- specifikacije artikala (red 'artikl' pa podaci ispod) ----------
    poz = L.nadji("artikl", stupac=1)
    if poz:
        r_hdr = poz[0]
        r = r_hdr + 1
        while r <= L.max_row:
            rb = _num(L.cell(r, 1))
            if rb is None:
                break
            dorade = [_txt(L.cell(r, cc)) for cc in (7, 8, 9, 10)]
            k.artikli.append(ArtiklSpec(
                rb=int(rb),
                sirina_mm=_num(L.cell(r, 2)),
                visina_mm=_num(L.cell(r, 3)),
                broj_boja=_num(L.cell(r, 4)),
                broj_prolaza=_num(L.cell(r, 5)),
                lak=_txt(L.cell(r, 6)),
                dorada=", ".join(d for d in dorade if d) or None,
                tip_papira=_txt(L.cell(r, 11)),
                gramatura=_txt(L.cell(r, 12)),
                nabavna_cijena=_txt(L.cell(r, 13)),
            ))
            r += 1

    # --- vanjska usluga iz tablice 'Proces' (stupac K) -------------------
    poz = L.nadji("Proces", stupac=1)
    if poz:
        r = poz[0] + 1
        while r <= L.max_row:
            a = L.cell(r, 1)
            if isinstance(a, str) and a.strip().lower().startswith("ukupni"):
                break
            if _num(a) is not None:
                k.vanjska_usluga += _num(L.cell(r, 11)) or 0.0
            r += 1

    # --- 'Proizvod' blok: nazivi artikala + udjeli -----------------------
    poz = L.nadji("Proizvod", stupac=1)
    if poz:
        r = poz[0] + 1
        i = 0
        while r <= L.max_row:
            e = _txt(L.cell(r, 5))
            if e and e.lower() == "total":
                break
            naziv = _txt(L.cell(r, 1))
            udio = _num(L.cell(r, 5))
            if naziv or udio is not None:
                if i < len(k.artikli):
                    k.artikli[i].naziv = naziv
                    k.artikli[i].udio = udio
                    k.artikli[i].kolicina = _num(L.cell(r, 6))
                elif naziv:
                    k.artikli.append(ArtiklSpec(naziv=naziv, udio=udio,
                                                kolicina=_num(L.cell(r, 6))))
                i += 1
            if naziv is None and udio is None and _txt(L.cell(r, 6)) is None:
                # potpuno prazan red — kraj bloka bez 'total' oznake
                break
            r += 1

    # --- glavni stroj ----------------------------------------------------
    poz = L.nadji("STROJ:", stupac=1)
    if poz:
        k.glavni_stroj = _txt(L.desno_od(*poz))

    # --- UTROŠAK MATERIJALA ----------------------------------------------
    poz = L.nadji("UTRO", stupac=1)  # "UTROŠAK MATERIJALA" (dijakritika varira)
    if poz is None:
        raise ValueError("Nije pronađen blok 'UTROŠAK MATERIJALA'.")
    r_mat = poz[0]
    poz_hdr = L.nadji("šifra", od_reda=r_mat) or L.nadji("ifra", od_reda=r_mat)
    if poz_hdr is None:
        raise ValueError("Nije pronađeno zaglavlje tablice materijala ('Šifra').")
    r = poz_hdr[0] + 1
    while r <= L.max_row:
        # kraj: red s 'UKUPNO' u stupcu J (10) ili bilo gdje u retku
        if any(isinstance(L.cell(r, cc), str) and "UKUPNO" in str(L.cell(r, cc)).upper()
               for cc in range(1, L.max_col + 1)):
            break
        sifra = _txt(L.cell(r, 1))
        naziv = _txt(L.cell(r, 3))
        if sifra or naziv:
            m = Materijal(
                sifra=sifra, naziv=naziv,
                jedinica=_txt(L.cell(r, 8)),
                kolicina=_num(L.cell(r, 9)),
                jed_cijena=_num(L.cell(r, 10)),
                vrijednost=_num(L.cell(r, 11)),
            )
            if m.vrijednost is None and m.kolicina is not None and m.jed_cijena is not None:
                m.vrijednost = m.kolicina * m.jed_cijena
            k.materijali.append(m)
        r += 1

    # --- UTROŠAK RADA ------------------------------------------------------
    poz = L.nadji("UTRO", stupac=1, od_reda=r)  # drugi UTROŠAK blok (rad)
    poz_hdr = L.nadji("Naziv STROJA", stupac=1)
    if poz_hdr is None:
        raise ValueError("Nije pronađeno zaglavlje tablice rada ('Naziv STROJA').")
    r = poz_hdr[0] + 1
    while r <= L.max_row:
        if any(isinstance(L.cell(r, cc), str) and "UKUPNO" in str(L.cell(r, cc)).upper()
               for cc in range(1, L.max_col + 1)):
            break
        stroj = _txt(L.cell(r, 1))
        operacija = _txt(L.cell(r, 3))
        if stroj or operacija:
            k.operacije.append(Operacija(
                stroj=stroj, operacija=operacija,
                proces=int(_num(L.cell(r, 7)) or 0) or None,
                vrijeme_min=_minute(L.cell(r, 9)),
                cijena_h=_num(L.cell(r, 10)),
                iznos=_num(L.cell(r, 11)),
            ))
        r += 1

    # --- PONOVNI IZRAČUN troškova (ne vjerujemo Excel kešu) ---------------
    k.trosak_materijal = round(sum(m.vrijednost or 0.0 for m in k.materijali), 4)
    k.trosak_rad = round(sum(o.iznos or 0.0 for o in k.operacije), 4)
    k.ck_total = round(k.trosak_materijal + k.trosak_rad + k.vanjska_usluga, 4)
    if k.serija:
        k.ck_po_jedinici = round(k.ck_total / k.serija, 6)

    # raspodjela CK po artiklima prema udjelu
    for a in k.artikli:
        if a.udio is not None:
            a.ck_total = round(a.udio * (k.trosak_materijal + k.trosak_rad), 4)

    # --- provjere / upozorenja ---------------------------------------------
    wb_f = openpyxl.load_workbook(putanja, data_only=False)
    dbl_eq = sum(1 for row in wb_f.active.iter_rows() for c in row
                 if isinstance(c.value, str) and c.value.startswith("=="))
    if dbl_eq:
        k.upozorenja.append(
            f"Izvorni Excel ima {dbl_eq} formula s '==' (slomljen izračun u izvoru) — "
            f"troškovi su ponovno izračunati iz stavki.")

    # sanity: usporedi naš zbroj materijala s Excelovim UKUPNO ako postoji
    if not k.materijali:
        k.upozorenja.append("Nije pronađen nijedan materijal!")
    if not k.operacije:
        k.upozorenja.append("Nije pronađena nijedna operacija rada!")
    if abs(sum(a.udio or 0 for a in k.artikli) - 1.0) > 0.02 and k.artikli:
        k.upozorenja.append("Zbroj udjela artikala nije 1 — provjeri raspodjelu.")

    return k
